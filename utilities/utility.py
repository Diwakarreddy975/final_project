import openpyxl

def read_data_from_excel(path,shett):
    data_list=[]
    wb=openpyxl.load_workbook(path)
    sh=wb[shett]
    row_count=sh.max_row
    col_count=sh.max_column
    for i in range(2,row_count+1):
        row=[]
        for j in range(1,col_count+1):
            row.append(sh.cell(row=i,column=j).value)
        data_list.append(row)
    return data_list
